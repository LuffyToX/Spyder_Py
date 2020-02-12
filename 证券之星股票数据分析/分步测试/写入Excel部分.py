# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import os

class Save_to_Excel:
    def __init__(self, title, data):
        self.title = title
        self.data = data

        # 新建工作簿
        self.wb = openpyxl.Workbook()
        # 获取活跃工作表
        self.sheet = self.wb.active


    def add_title(self):
        """ 向 Excel 中写入标题 """

        ''' 为当前工作表更名 '''
        self.sheet.title = "股市"

        sheet_list = self.wb.sheetnames
        print("当前工作簿拥有的工作表为：%s" % sheet_list)

        ''' 写入标题 '''
        self.sheet.append(self.title)
        print("**** 标题写入成功 ****")


    def add_data(self):
        """ 向 Excel 中写入数据 """

        code = list()
        simple = list()
        latest_price = list()
        ups_and_downs = list()
        ups_and_downs_val = list()
        ups_in_5mins = list()
        volume = list()
        busi_volume = list()
        turn_rate = list()
        amplitude = list()
        volume_ratio = list()
        comparison = list()
        market_rate = list()


        for i in range(0, len(self.data), 13):
            code.append(self.data[i])
            simple.append(self.data[i + 1])
            latest_price.append(self.data[i + 2])
            ups_and_downs.append(self.data[i + 3])
            ups_and_downs_val.append(self.data[i + 4])
            ups_in_5mins.append(self.data[i + 5])
            volume.append(self.data[i+6])
            busi_volume.append(self.data[i+7])
            turn_rate.append(self.data[i+8])
            amplitude.append(self.data[i+9])
            volume_ratio.append(self.data[i+10])
            comparison.append(self.data[i+11])
            market_rate.append(self.data[i+12])


        for num in range(13):
            col = get_column_letter(num + 1)  # 列号
            for row in range(len(self.data)//13):  # 行号
                col_and_row = col + str(row + 2)  # (要从第二行开始，否则会覆盖标题)
                if num == 0:
                    self.sheet[col_and_row] = code[row]
                elif num == 1:
                    self.sheet[col_and_row] = simple[row]
                elif num == 2:
                    self.sheet[col_and_row] = latest_price[row]
                elif num == 3:
                    self.sheet[col_and_row] = ups_and_downs[row]
                elif num == 4:
                    self.sheet[col_and_row] = ups_and_downs_val[row]
                elif num == 5:
                    self.sheet[col_and_row] = ups_in_5mins[row]
                elif num == 6:
                    self.sheet[col_and_row] = volume[row]
                elif num == 7:
                    self.sheet[col_and_row] = busi_volume[row]
                elif num == 8:
                    self.sheet[col_and_row] = turn_rate[row]
                elif num == 9:
                    self.sheet[col_and_row] = amplitude[row]
                elif num == 10:
                    self.sheet[col_and_row] = volume_ratio[row]
                elif num == 11:
                    self.sheet[col_and_row] = comparison[row]
                else:
                    self.sheet[col_and_row] = market_rate[row]

        print("**** 数据写入成功 ****")


    def set_width_and_height(self, width, height):
        """ 设置单元格行高和列宽 """

        # 列宽
        for num in range(13):
            col = get_column_letter(num + 1)
            self.sheet.column_dimensions[col].width = width
        # 行高
        self.sheet.row_dimensions[1].height = 20 # 标题行
        for row in range(len(self.data)//13): # 数据行
            self.sheet.row_dimensions[row+2].height = height


    def freeze_and_save(self):
        """ 冻结标题行以及保存 """

        # 如果要实现该函数，则需要写入标题和数据
        self.add_title()
        self.add_data()
        self.set_width_and_height(15, 15)

        self.sheet.freeze_panes = 'A2'

        self.wb.save('股市.xlsx')



if __name__ == "__main__":
    os.chdir(r'C:\Users\Dell\Desktop')
    title = ['代码', '简称', '最新价', '涨跌幅', '涨跌额', '5分钟涨幅',\
             '成交量(手)', '成交额(万元)', '换手率', '振幅', '量比', '委比', '市盈率']
    data = ['002867', '周大生', '33.72', '5.38%', '1.72', '-0.06%', '22015.08', '7346.65', '1.35%', '4.13%', '1.19', '-0.97', '32.31', '300379', '东方通', '20.00', '5.37%', '1.02', '0.30%', '86659.38', '17132.51', '4.34%', '8.80%', '1.29', '0.29', '444.84', '600230', '沧州大化', '17.08', '5.37%', '0.87', '0.35%', '166847.01', '28160.64', '4.05%', '5.37%', '0.86', '0.05', '2.96', '300346', '南大光电', '12.18', '5.36%', '0.62', '0.25%', '38312.40', '4567.03', '1.57%', '4.33%', '0.76', '-0.28', '195.57']
    Excel = Save_to_Excel(title, data)
    Excel.freeze_and_save()

